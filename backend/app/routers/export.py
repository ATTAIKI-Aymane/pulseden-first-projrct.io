from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session as DBSession
from fastapi.responses import Response
import pandas as pd
import io
import httpx

from app.database import get_db
from app import models

router = APIRouter(prefix="/sessions", tags=["Export"])


def build_export_data(session_id: int, db: DBSession):
    accounts = db.query(models.Account).filter(models.Account.session_id == session_id).all()
    if not accounts:
        raise HTTPException(status_code=400, detail="No accounts found for this session")

    rows = []
    for account in accounts:
        score = db.query(models.Score).filter(models.Score.account_id == account.id).first()
        sequence = db.query(models.Sequence).filter(models.Sequence.account_id == account.id).first()
        signals = db.query(models.Signal).filter(models.Signal.account_id == account.id).all()
        contact = db.query(models.Contact).filter(models.Contact.account_id == account.id).first()

        rows.append({
            "rank": score.rank if score else None,
            "company_name": account.company_name,
            "domain": account.domain,
            "industry": account.industry,
            "location": account.location,
            "fit_score": score.fit_score if score else None,
            "signal_score": score.signal_score if score else None,
            "total_score": score.total_score if score else None,
            "num_signals": len(signals),
            "top_signal": signals[0].description if signals else "",
            "contact_name": contact.full_name if contact else "",
            "contact_title": contact.job_title if contact else "",
            "contact_linkedin": contact.linkedin_url if contact else "",
            "contact_email": contact.email if contact else "",
            "outreach_subject": sequence.subject if sequence else "",
            "outreach_message": sequence.message_body if sequence else "",
        })

    rows.sort(key=lambda x: (x["rank"] is None, x["rank"]))
    return rows


@router.get("/{session_id}/export/csv")
def export_csv(session_id: int, db: DBSession = Depends(get_db)):
    data = build_export_data(session_id, db)
    df = pd.DataFrame(data)

    stream = io.StringIO()
    df.to_csv(stream, index=False)
    stream.seek(0)

    session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if session:
        export_record = models.Export(
            session_id=session_id,
            export_type="csv",
            destination="download",
            status="completed",
        )
        db.add(export_record)
        db.commit()

    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=pulsedev_export_session_{session_id}.csv"}
    )



@router.get("/{session_id}/export/excel")
def export_excel(session_id: int, db: DBSession = Depends(get_db)):
    data = build_export_data(session_id, db)
    df = pd.DataFrame(data)

    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Prospects")

        worksheet = writer.sheets["Prospects"]

        # Largeur des colonnes auto
        for i, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = min(max_len, 50)

        # Style header (bold + fond gris)
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Freeze la première ligne (headers toujours visibles)
        worksheet.freeze_panes = "A2"

    stream.seek(0)

    session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if session:
        export_record = models.Export(
            session_id=session_id,
            export_type="excel",
            destination="download",
            status="completed",
        )
        db.add(export_record)
        db.commit()

    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pulsedev_export_session_{session_id}.xlsx"}
    )


@router.post("/{session_id}/export/webhook")
def export_webhook(session_id: int, webhook_url: str, db: DBSession = Depends(get_db)):
    data = build_export_data(session_id, db)

    session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    status = "completed"
    try:
        with httpx.Client(timeout=10) as client:
            response = client.post(webhook_url, json={"session_id": session_id, "accounts": data})
            response.raise_for_status()
    except Exception as e:
        status = "failed"
        export_record = models.Export(
            session_id=session_id, export_type="webhook",
            destination=webhook_url, status=status,
        )
        db.add(export_record)
        db.commit()
        raise HTTPException(status_code=502, detail=f"Webhook delivery failed: {str(e)}")

    export_record = models.Export(
        session_id=session_id, export_type="webhook",
        destination=webhook_url, status=status,
    )
    db.add(export_record)

    session.status = "completed"
    db.commit()

    return {"message": f"Data exported successfully to webhook for {len(data)} accounts"}


@router.get("/{session_id}/export/preview")
def export_preview(session_id: int, db: DBSession = Depends(get_db)):
    """Aperçu JSON des données d'export (utile pour le frontend avant export réel)"""
    return build_export_data(session_id, db)